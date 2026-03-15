from openpyxl import load_workbook

def get_test_data(file, sheet):

    wb = load_workbook(file)
    sh = wb[sheet]

    data = []

    for row in sh.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data